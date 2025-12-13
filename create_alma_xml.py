from lxml import etree
from lxml.builder import ElementMaker
from lxml.etree import Comment
from openpyxl import load_workbook
import os
import datetime

def create_check_window (excel='results (2).xlsx', image_folder='2025_05_16_SHL-Images-Import/'):
    metadata_file = load_workbook(excel, data_only=True)
    metadata_all = []
    ws = metadata_file.active
    for item in ws.iter_rows(min_row = 2):
        item_data = [item[0].value, item[4].value, item[3].value, item[5].value]
        if item[1].value:
            item_data.append(item[1].value.replace("$h",""))
        else:
            item_data.append(item[6].value)
        metadata_all.append(item_data)
    metadata.update(find_item_metadata(image_folder, metadata_all))
    #metadata.sort(key=lambda sublist: sublist[2])
    get_metadata()

def get_metadata():
    return metadata

def find_item_metadata(image_folder, all_metadata):
    item_images = []
    for directory in os.listdir(image_folder):
        item_images.append(directory)
    excel_metadata = all_metadata
    new_metadata = {}
    for item in excel_metadata:
        if str(item[0]) in str(item_images):
            new_metadata[item[0]] = item
    print(new_metadata)
    return new_metadata
    
def create_xml(metadata, image_folder, xml_button):
    E = ElementMaker(namespace = "http://www.loc.gov/MARC21/slim", nsmap = {
        'marc' : 'http://www.loc.gov/MARC21/slim',
        'xsi' : 'http://www.w3.org/2001/XMLSchema-instance',
        'schemaLocation' : 'https://www.loc.gov/standards/marcxml/schema/MARC21slim.xsd'
    })
    
    new_feed = E.collection()
    file_num = 1
    for i in metadata.values():

        comment = Comment("New record starts here")
        new_feed.append(comment)

        record = E.record()
        new_feed.append(record)

        leader = E.leader(i[3])
        record.append(leader)

        mms_id = E.controlfield(i[2], {'tag' : '001'})
        record.append(mms_id)

        title = E.datafield({'tag' : '245'}, {'ind1' : '0'}, {'ind2' : '0'})
        record.append(title)
        title_sub = E.subfield(i[1], {'code' : 'a'}) 
        title.append(title_sub)

        collection_id = E.datafield({'tag' : '787'}, {'ind1' : '0'}, {'ind2' : ' '})
        record.append(collection_id)
        collection_id_sub = E.subfield('81325587200009266', {'code' : 'w'}) # to change to new digitised collections
        collection_id.append(collection_id_sub)
        image_num = 1
        for image_file in os.scandir(image_folder + '/' + str(i[0])):
            
            images = E.datafield({'tag' : '856'}, {'ind1' : '1'}, {'ind2' : '2'})

            record.append(images)
            images_sub = E.subfield(str(i[0]) + '/' + image_file.name, {'code' : 'u'})
            images.append(images_sub)
            images_sub = E.subfield(i[4], {'code' : 'z'})
            images.append(images_sub)
            if len(i)>5:
                images_sub = E.subfield(i[5], {'code' : 'n'})
            else:
                images_sub = E.subfield('Unrestricted', {'code' : 'n'})
            images.append(images_sub)
            #images_sub = E.subfield(str(file_num) + '.' + str(image_num) + "\\x", {'code' : '8'})
            #images.append(images_sub)
            #image_num += 1
        #image_num = 1
        file_num +=1
    file_num = 1

    file_name = image_folder + '/' + (datetime.datetime.now()).strftime('%Y_%m_%d_%H-%M-%S')+'_alma_import.xml'
    with open(file_name, 'wb') as f:
        f.write(etree.tostring(
            new_feed, encoding="utf-8", xml_declaration=False,
            pretty_print=True))
    xml_button.config(bg='#7fd1ae', text='Completed! Click to run again.')

global metadata
metadata = {}

if __name__ == "__main__":
    create_xml();